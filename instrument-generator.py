from pathlib import Path
import openpyxl
import orchestral_sections.makeGrid as makeGrid
import orchestral_sections.updateController as updateControllerSections
import orchestral_sections.makeSelector as makeSelector
import orchestral_doublers.makeSelector as makeSelectorDbl
import orchestral_doublers.OSCGrid as OSCGrid
import orchestral_doublers.updateController as updateController


# Retrieve excel info, and create working dict
# All functions receive a dict of instruments(strings) to work with


wb = openpyxl.load_workbook('./template-tracks-user.xlsx')
wss = wb.worksheets
sections = dict()
doublings = dict()
for i in range(1, len(wss)):
    ws = wss[i]
    if wb.sheetnames[i] == 'DOUBLING':
        doublings.update({ws.cell(row, 2).value: [ws.cell(row, col).value for col in range(
            3, ws.max_column+1) if ws.cell(row, col).value != None] for row in range(3, ws.max_row+1)})
        continue
    lrow = ws.max_row+1
    lcol = ws.max_column+1
    sections.update({wb.sheetnames[i]: {
        ws.cell(2, col).value: [
            ws.cell(row, col).value for row in range(3, lrow) if ws.cell(row, col).value != None
        ] for col in range(2, lcol)}})

# *Create cubase select macro
makeSelector.makeSelectors(sections)

# * Create and dump open stage control grid
makeGrid.make_grid(sections)

# * Updating midi controller
updateControllerSections.updateController(sections)

# *Generating cubase project logical editor xml's
makeSelectorDbl.generator(doublings)

# *Create and dump open stage control grid
OSCGrid.make_grid(doublings)

# *Update Midi controller
updateController.updateController(doublings)

print('Success!')
